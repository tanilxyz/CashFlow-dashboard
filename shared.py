from __future__ import annotations

import os
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils.cell import range_boundaries

DEFAULT_FILE = str(Path("data") / "Group MIS working file.v4.xlsx")

ENTITY_TABLES = {
    "A_INR": {
        "sheet": "A_INR",
        "opening_cash_row": 5,
        "currency": "INR",
        "ar": "ACO_AR",
        "ap": "ACO_AP",
        "opex": "ACO_OPEX",
        "cash_currency_label": "INR",
    },
    "E_SGD": {
        "sheet": "E_SGD",
        "opening_cash_row": 5,
        "currency": "SGD",
        "ar": "SCO_AR",
        "ap": "SCO_AP",
        "opex": "SCO_OPEX",
        "cash_currency_label": "SGD",
    },
}

ERROR_STRINGS = {"#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


# ============================================================
# Helpers
# ============================================================
def _is_nullish(x) -> bool:
    return x is None or pd.isna(x)


def _clean_scalar(x):
    if isinstance(x, str):
        s = x.strip()
        if s in ERROR_STRINGS or s == "":
            return np.nan
        return s
    return x


def normalize_col(col) -> str:
    if col is None:
        return ""
    s = str(col).strip().lower()
    s = s.replace("&", " and ")
    s = s.replace("/", "_")
    s = s.replace("?", "")
    s = s.replace("(", "").replace(")", "")
    s = s.replace("-", "_")
    s = s.replace("%", "pct")
    s = s.replace(".", "")
    s = s.replace(",", "")
    s = s.replace("  ", " ")
    s = s.replace(" ", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def first_non_null(df: pd.DataFrame, cols: List[str]) -> pd.Series:
    out = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    for c in cols:
        if c in df.columns:
            ser = pd.to_datetime(df[c], errors="coerce")
            out = out.fillna(ser)
    return out


def month_end(series: pd.Series) -> pd.Series:
    s = pd.to_datetime(series, errors="coerce")
    return s.dt.to_period("M").dt.to_timestamp("M")


def format_money(x, currency):
    if pd.isna(x):
        return ""
    if currency == "INR":
        return f"₹{x:,.0f}"
    return f"S${x:,.2f}"


# ============================================================
# Excel reading
# ============================================================
@st.cache_data(show_spinner=False)
def read_controls(excel_path: str) -> Dict[str, object]:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb["Controls"]
    return {
        "start_date": pd.to_datetime(ws["B3"].value),
        "weeks_horizon": int(ws["B4"].value or 13),
        "months_horizon": int(ws["B5"].value or 6),
        "fx_sgd_inr": float(ws["B6"].value or 72),
        "overdue_days": int(ws["B9"].value or 7),
        "large_exposure": float(ws["B10"].value or 5_000_000),
    }


@st.cache_data(show_spinner=False)
def read_opening_cash(excel_path: str, sheet_name: str) -> float:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    try:
        return float(ws["B5"].value or 0)
    except Exception:
        return 0.0


@st.cache_data(show_spinner=False)
def extract_table(excel_path: str, sheet_name: str, table_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    if table_name not in ws.tables:
        raise ValueError(f"Table {table_name} not found in sheet {sheet_name}.")

    ref = ws.tables[table_name].ref
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = list(
        ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    if not rows:
        return pd.DataFrame()

    headers = [normalize_col(h) for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=headers)
    df = df.apply(lambda col: col.map(_clean_scalar))
    df.columns = [normalize_col(c) for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def remove_repeated_header_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    cols = list(df.columns)

    def is_header_like(row) -> bool:
        vals = [row.get(c) for c in cols]
        non_null = [v for v in vals if not _is_nullish(v)]
        if len(non_null) < 2:
            return False
        hits = 0
        for c, v in zip(cols, vals):
            if _is_nullish(v):
                continue
            if normalize_col(v) == c:
                hits += 1
        return hits >= max(2, len(cols) // 2)

    mask = df.apply(is_header_like, axis=1)
    if mask.any():
        df = df.loc[~mask].copy()
    return df.reset_index(drop=True)


# ============================================================
# Cleaning / derived columns
# ============================================================
def add_derived_columns(df: pd.DataFrame, entity: str, section: str, start_date: pd.Timestamp) -> pd.DataFrame:
    df = df.copy()

    rename_map = {c: normalize_col(c) for c in df.columns}
    df = df.rename(columns=rename_map)

    for c in df.columns:
        if any(k in c for k in ["date", "month"]):
            df[c] = pd.to_datetime(df[c], errors="coerce")
        elif any(k in c for k in ["amount", "value", "pct", "week_no", "week"]):
            df[c] = pd.to_numeric(df[c], errors="coerce")

    if section == "ar":
        df["entity"] = entity
        df["section"] = "AR"
        df["invoice_month_end"] = month_end(df["invoice_date"]) if "invoice_date" in df.columns else pd.NaT
        df["scf_month_end"] = month_end(df["scf_date"]) if "scf_date" in df.columns else pd.NaT

        invoice_value = pd.to_numeric(df.get("invoice_value"), errors="coerce") if "invoice_value" in df.columns else pd.Series(np.nan, index=df.index)
        scf_pct = pd.to_numeric(df.get("scf_fee_pct"), errors="coerce") if "scf_fee_pct" in df.columns else pd.Series(0.0, index=df.index)
        scf_flag = df.get("scf", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()
        scf_yes = scf_flag.eq("yes")

        df["scf_fee_amount_calc"] = np.where(scf_yes, invoice_value * scf_pct, 0.0)
        if "scf_fee_amount" in df.columns:
            df["scf_fee_amount_clean"] = pd.to_numeric(df["scf_fee_amount"], errors="coerce").fillna(df["scf_fee_amount_calc"])
        else:
            df["scf_fee_amount_clean"] = df["scf_fee_amount_calc"]

        if "net_receipt_amount" in df.columns:
            df["net_receipt_amount_clean"] = pd.to_numeric(df["net_receipt_amount"], errors="coerce")
        else:
            df["net_receipt_amount_clean"] = np.nan
        df["net_receipt_amount_clean"] = df["net_receipt_amount_clean"].fillna(invoice_value - df["scf_fee_amount_clean"])

        manual = pd.to_datetime(df.get("manual_expected_receipt_date"), errors="coerce") if "manual_expected_receipt_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        linked = pd.to_datetime(df.get("linked_expected_receipt_date"), errors="coerce") if "linked_expected_receipt_date" in df.columns else manual
        receipt_date = pd.to_datetime(df.get("receipt_date"), errors="coerce") if "receipt_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        scf_date = pd.to_datetime(df.get("scf_date"), errors="coerce") if "scf_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        receipt_status = df.get("receipt_status", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()

        cash_date = linked.copy()
        cash_date = cash_date.where(~receipt_status.eq("received"), receipt_date)
        cash_date = cash_date.where(~scf_yes, scf_date)
        cash_date = cash_date.fillna(linked).fillna(manual).fillna(receipt_date).fillna(scf_date)
        df["cash_date"] = cash_date
        df["week_no"] = np.floor((df["cash_date"] - start_date).dt.days / 7).astype("Int64") + 1

    elif section == "ap":
        df["entity"] = entity
        df["section"] = "AP"
        df["invoice_month_end"] = month_end(df["invoice_date"]) if "invoice_date" in df.columns else pd.NaT

        amount_payable = pd.to_numeric(df.get("amount_payable"), errors="coerce") if "amount_payable" in df.columns else pd.Series(np.nan, index=df.index)
        gst_itc = pd.to_numeric(df.get("gst_itc_amount"), errors="coerce") if "gst_itc_amount" in df.columns else pd.Series(0.0, index=df.index)
        net_amount = pd.to_numeric(df.get("net_amount"), errors="coerce") if "net_amount" in df.columns else pd.Series(np.nan, index=df.index)

        df["amount_payable_clean"] = amount_payable
        derived_net = (amount_payable - gst_itc).clip(lower=0)
        if net_amount.notna().any():
            df["net_amount_clean"] = net_amount.fillna(derived_net)
        else:
            df["net_amount_clean"] = derived_net
        df["amount_for_pnl"] = df["net_amount_clean"].fillna(amount_payable)

        payment_status = df.get("payment_status", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()
        payment_date = pd.to_datetime(df.get("payment_date"), errors="coerce") if "payment_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        due_date = pd.to_datetime(df.get("due_date"), errors="coerce") if "due_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        cash_date = due_date.copy()
        cash_date = cash_date.where(~payment_status.eq("paid"), payment_date)
        cash_date = cash_date.fillna(payment_date).fillna(due_date)
        df["cash_date"] = cash_date
        df["week_no"] = np.floor((df["cash_date"] - start_date).dt.days / 7).astype("Int64") + 1

    else:
        df["entity"] = entity
        df["section"] = "OPEX"
        df["invoice_month_end"] = month_end(df["invoice_date"]) if "invoice_date" in df.columns else pd.NaT

        amount = pd.to_numeric(df.get("amount"), errors="coerce") if "amount" in df.columns else pd.Series(np.nan, index=df.index)
        gst_itc = pd.to_numeric(df.get("gst_itc_amount"), errors="coerce") if "gst_itc_amount" in df.columns else pd.Series(0.0, index=df.index)
        net_amount = pd.to_numeric(df.get("net_amount"), errors="coerce") if "net_amount" in df.columns else pd.Series(np.nan, index=df.index)

        df["amount_clean"] = amount
        derived_net = (amount - gst_itc).clip(lower=0)
        if net_amount.notna().any():
            df["net_amount_clean"] = net_amount.fillna(derived_net)
        else:
            df["net_amount_clean"] = derived_net
        df["amount_for_pnl"] = df["net_amount_clean"].fillna(amount)

        status = df.get("status", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()
        payment_date = pd.to_datetime(df.get("payment_date"), errors="coerce") if "payment_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        due_date = pd.to_datetime(df.get("due_date"), errors="coerce") if "due_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        cash_date = due_date.copy()
        cash_date = cash_date.where(~status.eq("paid"), payment_date)
        cash_date = cash_date.fillna(payment_date).fillna(due_date)
        df["cash_date"] = cash_date
        df["week_no"] = np.floor((df["cash_date"] - start_date).dt.days / 7).astype("Int64") + 1

    if "week_no" in df.columns:
        df["week_no"] = pd.to_numeric(df["week_no"], errors="coerce").astype("Int64")
    if "invoice_month_end" in df.columns:
        df["invoice_month_end"] = pd.to_datetime(df["invoice_month_end"], errors="coerce")

    return df


def clean_section_table(df: pd.DataFrame, entity: str, section: str, start_date: pd.Timestamp) -> pd.DataFrame:
    if df.empty:
        return df

    df = remove_repeated_header_rows(df)
    df = add_derived_columns(df, entity=entity, section=section, start_date=start_date)

    essential = {
        "ar": ["invoice_date", "invoice_value", "net_receipt_amount_clean", "cash_date"],
        "ap": ["invoice_date", "amount_payable_clean", "amount_for_pnl", "cash_date"],
        "opex": ["invoice_date", "amount_clean", "amount_for_pnl", "cash_date"],
    }[section]

    mask = pd.Series(False, index=df.index)
    for c in essential:
        if c in df.columns:
            mask = mask | df[c].notna()
    df = df.loc[mask].copy()

    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].replace(list(ERROR_STRINGS), np.nan)

    return df.reset_index(drop=True)


# ============================================================
# Forecast builders
# ============================================================
def build_weekly_forecast(
    ar: pd.DataFrame,
    ap: pd.DataFrame,
    opex: pd.DataFrame,
    opening_cash: float,
    start_date: pd.Timestamp,
    weeks_horizon: int,
    currency: str,
) -> pd.DataFrame:
    weeks = pd.DataFrame({"week_no": range(1, weeks_horizon + 1)})
    weeks["week_start"] = start_date + pd.to_timedelta((weeks["week_no"] - 1) * 7, unit="D")
    weeks["week_end"] = weeks["week_start"] + pd.to_timedelta(6, unit="D")

    ar_f = ar.loc[ar["week_no"].between(1, weeks_horizon, inclusive="both")].copy() if not ar.empty else ar.copy()
    ap_f = ap.loc[ap["week_no"].between(1, weeks_horizon, inclusive="both")].copy() if not ap.empty else ap.copy()
    opex_f = opex.loc[opex["week_no"].between(1, weeks_horizon, inclusive="both")].copy() if not opex.empty else opex.copy()

    if ar_f.empty:
        ar_grp = pd.Series(dtype=float)
        scf_grp = pd.Series(dtype=float)
    else:
        ar_amt = "net_receipt_amount_clean" if "net_receipt_amount_clean" in ar_f.columns else "invoice_value"
        ar_grp = ar_f.groupby("week_no")[ar_amt].sum(min_count=1)
        scf_grp = ar_f.groupby("week_no")["scf_fee_amount_clean"].sum(min_count=1) if "scf_fee_amount_clean" in ar_f.columns else pd.Series(dtype=float)

    if ap_f.empty:
        ap_grp = pd.Series(dtype=float)
    else:
        ap_col = "amount_payable_clean" if "amount_payable_clean" in ap_f.columns else "amount_for_pnl"
        ap_grp = ap_f.groupby("week_no")[ap_col].sum(min_count=1)

    if opex_f.empty:
        opex_grp = pd.Series(dtype=float)
    else:
        op_col = "amount_clean" if "amount_clean" in opex_f.columns else "amount_for_pnl"
        opex_grp = opex_f.groupby("week_no")[op_col].sum(min_count=1)

    weeks["inflows"] = weeks["week_no"].map(ar_grp).fillna(0.0)
    weeks["ap_outflows"] = weeks["week_no"].map(ap_grp).fillna(0.0)
    weeks["opex_outflows"] = weeks["week_no"].map(opex_grp).fillna(0.0)
    weeks["scf_fee"] = weeks["week_no"].map(scf_grp).fillna(0.0)

    weeks["outflows"] = weeks["ap_outflows"] + weeks["opex_outflows"]
    weeks["net"] = weeks["inflows"] - weeks["outflows"]
    weeks["closing_cash"] = opening_cash + weeks["net"].cumsum()
    weeks["week_label"] = weeks.apply(lambda r: f"W{int(r['week_no'])} ({r['week_start']:%d-%b} to {r['week_end']:%d-%b})", axis=1)
    weeks["currency"] = currency
    weeks["opening_cash"] = opening_cash
    return weeks


def build_monthly_pnl(
    ar: pd.DataFrame,
    ap: pd.DataFrame,
    opex: pd.DataFrame,
    start_date: pd.Timestamp,
    months_horizon: int,
    currency: str,
) -> pd.DataFrame:
    month_ends = pd.date_range(start=start_date, periods=months_horizon, freq="ME")
    pnl = pd.DataFrame({"month_end": month_ends})
    pnl["month_label"] = pnl["month_end"].dt.strftime("%b-%Y")

    ar = ar.copy()
    ap = ap.copy()
    opex = opex.copy()
    ar["invoice_month_end"] = pd.to_datetime(ar.get("invoice_month_end"), errors="coerce")
    ap["invoice_month_end"] = pd.to_datetime(ap.get("invoice_month_end"), errors="coerce")
    opex["invoice_month_end"] = pd.to_datetime(opex.get("invoice_month_end"), errors="coerce")

    rev = (
        ar.loc[ar["invoice_month_end"].isin(month_ends)]
        .groupby("invoice_month_end")["invoice_value" if "invoice_value" in ar.columns else "amount_for_pnl"]
        .sum(min_count=1)
    )

    cogs = pd.Series(dtype=float)
    if not ap.empty:
        ap_cogs = ap.copy()
        if "category" in ap_cogs.columns:
            ap_cogs = ap_cogs.loc[ap_cogs["category"].astype(str).str.upper().eq("COGS")]
        ap_amt_col = "amount_for_pnl" if "amount_for_pnl" in ap_cogs.columns else "amount_payable_clean"
        cogs = (
            ap_cogs.loc[ap_cogs["invoice_month_end"].isin(month_ends)]
            .groupby("invoice_month_end")[ap_amt_col]
            .sum(min_count=1)
        )

    opex_amt_col = "amount_for_pnl" if "amount_for_pnl" in opex.columns else "amount_clean"
    opex_sum = (
        opex.loc[opex["invoice_month_end"].isin(month_ends)]
        .groupby("invoice_month_end")[opex_amt_col]
        .sum(min_count=1)
    )

    scf = pd.Series(dtype=float)
    if "scf_month_end" in ar.columns:
        scf_col = "scf_fee_amount_clean" if "scf_fee_amount_clean" in ar.columns else "scf_fee_amount"
        scf = (
            ar.loc[ar["scf_month_end"].isin(month_ends)]
            .groupby("scf_month_end")[scf_col]
            .sum(min_count=1)
        )

    pnl["revenue"] = pnl["month_end"].map(rev).fillna(0.0)
    pnl["cogs"] = pnl["month_end"].map(cogs).fillna(0.0)
    pnl["gross_profit"] = pnl["revenue"] - pnl["cogs"]
    pnl["opex"] = pnl["month_end"].map(opex_sum).fillna(0.0)
    pnl["scf_finance_cost"] = pnl["month_end"].map(scf).fillna(0.0)
    pnl["profit_loss"] = pnl["gross_profit"] - pnl["opex"] - pnl["scf_finance_cost"]
    pnl["ebitda"] = pnl["profit_loss"]
    pnl["currency"] = currency
    return pnl


def build_overview_metrics(
    ar: pd.DataFrame,
    ap: pd.DataFrame,
    opex: pd.DataFrame,
    weekly: pd.DataFrame,
    pnl: pd.DataFrame,
    start_date: pd.Timestamp,
    overdue_days: int,
    large_exposure: float,
) -> Dict[str, float]:
    ar_open = ar.copy()
    if "receipt_status" in ar_open.columns:
        open_mask = ~ar_open["receipt_status"].astype(str).str.lower().eq("received")
        received_mask = ar_open["receipt_status"].astype(str).str.lower().eq("received")
    else:
        open_mask = ar_open["cash_date"].notna()
        received_mask = ar_open["cash_date"].notna()

    overdue_mask = open_mask & ar_open["cash_date"].notna() & (ar_open["cash_date"] < (start_date - pd.Timedelta(days=overdue_days)))
    missing_ar_cash = int(ar_open["cash_date"].isna().sum()) if "cash_date" in ar_open.columns else 0
    missing_ap_cash = int(ap["cash_date"].isna().sum()) if "cash_date" in ap.columns else 0

    total_inflows = float(weekly["inflows"].sum()) if not weekly.empty else 0.0
    total_outflows = float(weekly["outflows"].sum()) if not weekly.empty else 0.0
    net_movement = float(weekly["net"].sum()) if not weekly.empty else 0.0
    closing_cash = float(weekly["closing_cash"].iloc[-1]) if not weekly.empty else 0.0
    negative_weeks = int((weekly["closing_cash"] < 0).sum()) if not weekly.empty else 0

    ar_base = pd.to_numeric(ar.get("invoice_value", ar.get("net_receipt_amount_clean", 0)), errors="coerce").fillna(0)
    ar_received_base = pd.to_numeric(ar.get("net_receipt_amount_clean", ar.get("invoice_value", 0)), errors="coerce").fillna(0)
    ap_base = pd.to_numeric(ap.get("amount_payable_clean", ap.get("amount_for_pnl", 0)), errors="coerce").fillna(0)
    ap_paid_base = pd.to_numeric(ap.get("amount_for_pnl", ap.get("amount_payable_clean", 0)), errors="coerce").fillna(0)

    ar_total_amount = float(ar_base.sum()) if len(ar_base) else 0.0
    ar_received_amount = float(ar_received_base.loc[received_mask].sum()) if len(ar_received_base) else 0.0
    ap_total_amount = float(ap_base.sum()) if len(ap_base) else 0.0
    if "payment_status" in ap.columns:
        paid_mask = ap["payment_status"].astype(str).str.lower().eq("paid")
    else:
        paid_mask = ap["cash_date"].notna()
    ap_paid_amount = float(ap_paid_base.loc[paid_mask].sum()) if len(ap_paid_base) else 0.0

    ar_received_pct = (ar_received_amount / ar_total_amount * 100.0) if ar_total_amount else 0.0
    ap_paid_pct = (ap_paid_amount / ap_total_amount * 100.0) if ap_total_amount else 0.0

    large_ar = int(ar.loc[pd.to_numeric(ar.get("net_receipt_amount_clean", ar.get("invoice_value", 0)), errors="coerce").fillna(0).abs() >= large_exposure].shape[0])
    large_ap = int(ap.loc[pd.to_numeric(ap.get("amount_for_pnl", ap.get("amount_payable_clean", 0)), errors="coerce").fillna(0).abs() >= large_exposure].shape[0])
    large_opex = int(opex.loc[pd.to_numeric(opex.get("amount_for_pnl", opex.get("amount_clean", 0)), errors="coerce").fillna(0).abs() >= large_exposure].shape[0])

    profit_loss_total = float(pnl["profit_loss"].sum()) if not pnl.empty and "profit_loss" in pnl.columns else float(pnl["ebitda"].sum()) if not pnl.empty and "ebitda" in pnl.columns else 0.0

    return {
        "total_inflows": total_inflows,
        "total_outflows": total_outflows,
        "net_movement": net_movement,
        "closing_cash": closing_cash,
        "negative_weeks": negative_weeks,
        "overdue_ar": int(overdue_mask.sum()),
        "missing_ar_cash": missing_ar_cash,
        "missing_ap_cash": missing_ap_cash,
        "large_ar": large_ar,
        "large_ap": large_ap,
        "large_opex": large_opex,
        "ar_total_amount": ar_total_amount,
        "ar_received_amount": ar_received_amount,
        "ap_total_amount": ap_total_amount,
        "ap_paid_amount": ap_paid_amount,
        "ar_received_pct": ar_received_pct,
        "ap_paid_pct": ap_paid_pct,
        "profit_loss_total": profit_loss_total,
    }


# ============================================================
# Rendering
# ============================================================
def money_text(x, currency):
    return format_money(x, currency)


def apply_visual_theme():
    st.markdown(
        """
        <style>
            .stApp {
                background: #232845;
                color: white;
            }

            .block-container {
                padding-top: 1rem;
                padding-bottom: 1.5rem;
            }

            .page-title {
                font-size: 2.15rem;
                font-weight: 800;
                color: white;
                margin-bottom: 0.25rem;
            }

            .page-subtitle {
                color: #d5d9ea;
                font-size: 0.95rem;
                line-height: 1.35;
                margin-bottom: 1rem;
                max-width: 1200px;
            }

            .metric-card {
                background: #2a3051;
                border: 1px solid rgba(255,255,255,0.16);
                border-radius: 4px;
                overflow: hidden;
                min-height: 116px;
            }

            .metric-head {
                background: white;
                color: #1f2431;
                font-size: 0.92rem;
                font-weight: 800;
                text-align: center;
                padding: 6px 8px;
                border-bottom: 1px solid rgba(0,0,0,0.08);
            }

            .metric-body {
                height: 76px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 1.75rem;
                font-weight: 800;
                color: white;
                padding: 0 8px;
                text-align: center;
            }

            .section-title {
                color: white;
                font-size: 1.15rem;
                font-weight: 800;
                text-align: center;
                margin: 0.75rem 0 0.35rem 0;
            }

            .small-note {
                color: #b8bdd6;
                font-size: 0.82rem;
            }

            div[data-testid="stDataFrame"] {
                border: 1px solid rgba(255,255,255,0.12);
                border-radius: 6px;
                overflow: hidden;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def metric_card(title: str, value, currency: Optional[str] = None, is_number: bool = True) -> None:
    if is_number and currency:
        value_text = money_text(value, currency)
    else:
        value_text = str(value)

    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-head">{title}</div>
            <div class="metric-body">{value_text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def aggregate_weekly_to_monthly_cash(weekly: pd.DataFrame) -> pd.DataFrame:
    if weekly.empty:
        return pd.DataFrame(columns=["month_end", "month_label", "cash_going_on", "cash_going_out", "ending_cash_on_hand", "profit_loss"])

    df = weekly.copy()
    df["month_end"] = pd.to_datetime(df["week_start"], errors="coerce").dt.to_period("M").dt.to_timestamp("M")
    grouped = (
        df.groupby("month_end", as_index=False)
        .agg(
            inflows=("inflows", "sum"),
            ap_outflows=("ap_outflows", "sum"),
            opex_outflows=("opex_outflows", "sum"),
            net=("net", "sum"),
            closing_cash=("closing_cash", "last"),
            opening_cash=("opening_cash", "first"),
        )
        .sort_values("month_end")
    )
    grouped["cash_going_on"] = grouped["inflows"]
    grouped["cash_going_out"] = grouped["ap_outflows"] + grouped["opex_outflows"]
    grouped["ending_cash_on_hand"] = grouped["closing_cash"]
    grouped["profit_loss"] = grouped["net"]
    grouped["month_label"] = grouped["month_end"].dt.strftime("%b")
    return grouped


def make_cash_flow_chart(monthly_cash: pd.DataFrame, currency: str) -> go.Figure:
    fig = go.Figure()
    fig.add_bar(
        x=monthly_cash["month_label"],
        y=monthly_cash["cash_going_on"],
        name="Cash going on",
        marker_color="#ff7a18",
    )
    fig.add_bar(
        x=monthly_cash["month_label"],
        y=monthly_cash["cash_going_out"],
        name="Cash going out",
        marker_color="#ff4fa3",
    )
    fig.add_scatter(
        x=monthly_cash["month_label"],
        y=monthly_cash["ending_cash_on_hand"],
        name="Ending Cash on Hand",
        mode="lines+markers",
        line=dict(color="white", width=2.5),
        marker=dict(symbol="square", size=8, color="white"),
    )

    fig.update_layout(
        title=dict(text="<b>Cash Flow</b>", x=0.5, xanchor="center"),
        barmode="group",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white"),
        margin=dict(l=25, r=25, t=55, b=25),
        height=360,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        xaxis=dict(showgrid=False),
        yaxis=dict(gridcolor="rgba(255,255,255,0.10)", title=f"Amount ({currency})"),
    )
    return fig


def make_profit_loss_chart(pnl: pd.DataFrame, currency: str) -> go.Figure:
    if pnl.empty:
        fig = go.Figure()
    else:
        colors = ["#ffb07c" if v >= 0 else "#ff4fa3" for v in pnl["profit_loss"]]
        fig = go.Figure()
        fig.add_bar(
            x=pnl["month_label"],
            y=pnl["profit_loss"],
            marker_color=colors,
            name="Profit / Loss",
        )
        fig.add_hline(y=0, line_width=1, line_color="rgba(255,255,255,0.45)")

    fig.update_layout(
        title=dict(text="<b>Profit / Loss</b>", x=0.5, xanchor="center"),
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white"),
        margin=dict(l=25, r=25, t=55, b=25),
        height=330,
        xaxis=dict(showgrid=False),
        yaxis=dict(gridcolor="rgba(255,255,255,0.10)", title=f"Amount ({currency})"),
    )
    return fig


def make_donut_chart(title: str, amount, pct: float, center_label: str, fill_color: str) -> go.Figure:
    pct = max(0, min(100, float(pct)))
    fig = go.Figure(
        data=[
            go.Pie(
                values=[pct, 100 - pct],
                hole=0.72,
                sort=False,
                direction="clockwise",
                rotation=90,
                marker=dict(colors=[fill_color, "#2f3557"]),
                textinfo="none",
                hoverinfo="skip",
                showlegend=False,
            )
        ]
    )

    fig.add_annotation(
        x=0.5,
        y=0.63,
        text=f"<b>{amount:,.2f}</b>" if isinstance(amount, (int, float, np.integer, np.floating)) else f"<b>{amount}</b>",
        showarrow=False,
        font=dict(color="white", size=22),
    )
    fig.add_annotation(
        x=0.5,
        y=0.40,
        text=f"<b>{pct:.0f}%</b><br>{center_label}",
        showarrow=False,
        font=dict(color="white", size=14),
    )

    fig.update_layout(
        title=dict(text=f"<b>{title}</b>", x=0.5, xanchor="center"),
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=55, b=20),
        height=330,
    )
    return fig


def render_dashboard(entity: str, data: Dict[str, pd.DataFrame]):
    currency = ENTITY_TABLES[entity]["currency"]
    display_title = "Cash Flow Dashboard" if entity == "A_INR" else "Cash Flow Dashboard"

    apply_visual_theme()

    st.markdown(f"<div class='page-title'>{display_title}</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class="page-subtitle">
            Following slide covers cash flow dashboard covering details like beginning cash on hand,
            cash going in, cash going out, profit/loss and ending cash on hand.
            It also includes detail of accounts receivable and payable.
        </div>
        """,
        unsafe_allow_html=True,
    )

    opening_cash = data.get("opening_cash", 0.0)
    metrics = data["metrics"]
    pnl = data["pnl"]
    weekly = data["weekly"]

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        metric_card("Beginning Cash on Hand", opening_cash, currency=currency)
    with c2:
        metric_card("Cash going on", metrics["total_inflows"], currency=currency)
    with c3:
        metric_card("Cash going out", metrics["total_outflows"], currency=currency)
    with c4:
        metric_card("Profit / Loss", metrics["profit_loss_total"], currency=currency)
    with c5:
        metric_card("Ending Cash on Hand", metrics["closing_cash"], currency=currency)

    st.write("")
    monthly_cash = aggregate_weekly_to_monthly_cash(weekly)
    st.plotly_chart(make_cash_flow_chart(monthly_cash, currency), use_container_width=True)

    l1, l2, l3 = st.columns([1.25, 1, 1])
    with l1:
        st.plotly_chart(make_profit_loss_chart(pnl, currency), use_container_width=True)
    with l2:
        st.plotly_chart(
            make_donut_chart(
                "Accounts Receivable",
                metrics["ar_received_amount"],
                metrics["ar_received_pct"],
                "Received",
                "#ff4fa3",
            ),
            use_container_width=True,
        )
    with l3:
        st.plotly_chart(
            make_donut_chart(
                "Accounts Payable",
                metrics["ap_paid_amount"],
                metrics["ap_paid_pct"],
                "Paid",
                "#ff9a56",
            ),
            use_container_width=True,
        )

    st.write("")
    st.subheader("Weekly cash forecast")
    weekly_disp = weekly[["week_no", "week_start", "week_end", "inflows", "ap_outflows", "opex_outflows", "net", "closing_cash"]].copy()
    weekly_disp.columns = ["Week", "Week Start", "Week End", "Inflows", "AP Outflows", "OPEX Outflows", "Net", "Closing Cash"]
    st.dataframe(
        weekly_disp.style.format(
            {
                "Inflows": lambda x: format_money(x, currency),
                "AP Outflows": lambda x: format_money(x, currency),
                "OPEX Outflows": lambda x: format_money(x, currency),
                "Net": lambda x: format_money(x, currency),
                "Closing Cash": lambda x: format_money(x, currency),
            }
        ),
        use_container_width=True,
        height=300,
    )

    st.subheader("Monthly P&L projection")
    pnl_disp = pnl[["month_label", "revenue", "cogs", "gross_profit", "opex", "scf_finance_cost", "profit_loss"]].copy()
    pnl_disp.columns = ["Month", "Revenue", "COGS", "Gross Profit", "OPEX", "SCF Finance Cost", "Profit / Loss"]
    st.dataframe(
        pnl_disp.style.format(
            {
                "Revenue": lambda x: format_money(x, currency),
                "COGS": lambda x: format_money(x, currency),
                "Gross Profit": lambda x: format_money(x, currency),
                "OPEX": lambda x: format_money(x, currency),
                "SCF Finance Cost": lambda x: format_money(x, currency),
                "Profit / Loss": lambda x: format_money(x, currency),
            }
        ),
        use_container_width=True,
        height=300,
    )

    with st.expander("Cleaned source tables"):
        st.write("AR")
        st.dataframe(data["ar"], use_container_width=True, height=240)
        st.write("AP")
        st.dataframe(data["ap"], use_container_width=True, height=240)
        st.write("OPEX")
        st.dataframe(data["opex"], use_container_width=True, height=240)

    st.download_button(
        "Download weekly forecast CSV",
        weekly.to_csv(index=False).encode("utf-8"),
        file_name=f"{entity}_weekly_forecast.csv",
        mime="text/csv",
    )
    st.download_button(
        "Download monthly P&L CSV",
        pnl.to_csv(index=False).encode("utf-8"),
        file_name=f"{entity}_monthly_pnl.csv",
        mime="text/csv",
    )


# ============================================================
# Pipeline
# ============================================================
@st.cache_data(show_spinner=False)
def load_entity_dashboard(excel_path: str, entity: str, mtime: float) -> Dict[str, pd.DataFrame]:
    if entity not in ENTITY_TABLES:
        raise ValueError(f"Unknown entity: {entity}")

    cfg = ENTITY_TABLES[entity]
    controls = read_controls(excel_path)
    start_date = pd.to_datetime(controls["start_date"])
    opening_cash = read_opening_cash(excel_path, cfg["sheet"])

    ar_raw = extract_table(excel_path, cfg["sheet"], cfg["ar"])
    ap_raw = extract_table(excel_path, cfg["sheet"], cfg["ap"])
    opex_raw = extract_table(excel_path, cfg["sheet"], cfg["opex"])

    ar = clean_section_table(ar_raw, entity, "ar", start_date)
    ap = clean_section_table(ap_raw, entity, "ap", start_date)
    opex = clean_section_table(opex_raw, entity, "opex", start_date)

    weekly = build_weekly_forecast(ar, ap, opex, opening_cash, start_date, controls["weeks_horizon"], cfg["currency"])
    pnl = build_monthly_pnl(ar, ap, opex, start_date, controls["months_horizon"], cfg["currency"])
    metrics = build_overview_metrics(ar, ap, opex, weekly, pnl, start_date, controls["overdue_days"], controls["large_exposure"])

    return {
        "ar": ar,
        "ap": ap,
        "opex": opex,
        "weekly": weekly,
        "pnl": pnl,
        "metrics": metrics,
        "controls": controls,
        "opening_cash": opening_cash,
    }


def run_app(entity_fixed: Optional[str] = None, default_file: str = DEFAULT_FILE):
    st.set_page_config(page_title="MIS Dashboard", layout="wide")
    st.sidebar.title("Settings")

    uploaded = st.sidebar.file_uploader("Upload workbook", type=["xlsx"])
    if uploaded is not None:
        tmp_path = Path("/tmp/uploaded_mis.xlsx")
        tmp_path.write_bytes(uploaded.read())
        excel_path = str(tmp_path)
    else:
        excel_path = default_file

    if not Path(excel_path).exists():
        st.info("Upload the Excel workbook or place it next to this app as 'Group MIS working file.v4.xlsx'.")
        return

    controls = read_controls(excel_path)
    st.sidebar.write("Start date:", controls["start_date"].date())
    st.sidebar.write("Weeks horizon:", controls["weeks_horizon"])
    st.sidebar.write("Months horizon:", controls["months_horizon"])
    st.sidebar.write("FX SGD/INR:", controls["fx_sgd_inr"])

    if entity_fixed in ENTITY_TABLES:
        entity = entity_fixed
    else:
        entity = st.sidebar.radio("Dashboard", ["A_INR", "E_SGD"], index=0)

    st.sidebar.write(f"Opening cash ({entity}): {read_opening_cash(excel_path, ENTITY_TABLES[entity]['sheet']):,.2f}")

    with st.sidebar:
        if st.button("🔄 Refresh Data"):
            st.cache_data.clear()
            st.rerun()

    last_modified = os.path.getmtime(excel_path)
    st.sidebar.caption(f"Last updated: {pd.Timestamp.fromtimestamp(last_modified):%d %b %Y, %H:%M:%S}")

    with st.spinner(f"Building {entity} dashboard..."):
        data = load_entity_dashboard(excel_path, entity, last_modified)

    render_dashboard(entity, data)
    st.caption(
        "This app reads the Excel Tables directly from the workbook, cleans the AR/AP/OPEX sections, and rebuilds the forecast logic inside Streamlit."
    )
