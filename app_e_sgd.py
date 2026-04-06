
import math
import os
import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils.cell import range_boundaries


# ==================    ==========================================
# CONFIG
# ============================================================
DEFAULT_FILE = "data\Group MIS working file.v4.xlsx"

# If you want two separate standalone dashboards, copy this file
# twice and set ENTITY_FIXED to "A_INR" or "E_SGD".
ENTITY_FIXED = "E_SGD"


ENTITY_TABLES = {
    "A_INR": {
        "sheet": "A_INR",
        "opening_cash_row": 5,
        "currency": "INR",
        "ar": "ACO_AR",
        "ap": "ACO_AP",
        "opex": "ACO_OPEX",
        "cash_currency_label": "INR",
    },
    "E_SGD": {
        "sheet": "E_SGD",
        "opening_cash_row": 5,
        "currency": "SGD",
        "ar": "SCO_AR",
        "ap": "SCO_AP",
        "opex": "SCO_OPEX",
        "cash_currency_label": "SGD",
    },
}

ERROR_STRINGS = {
    "#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"
}


# ============================================================
# HELPERS
# ============================================================
def _is_nullish(x) -> bool:
    return x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x)


def _clean_scalar(x):
    if isinstance(x, str):
        s = x.strip()
        if s in ERROR_STRINGS:
            return np.nan
        if s == "":
            return np.nan
        return s
    return x


def normalize_col(col) -> str:
    if col is None:
        return ""
    s = str(col).strip().lower()
    s = s.replace("&", " and ")
    s = s.replace("/", "_")
    s = s.replace("?", "")
    s = s.replace("(", "").replace(")", "")
    s = s.replace("-", "_")
    s = s.replace("%", "pct")
    s = s.replace(".", "")
    s = s.replace(",", "")
    s = s.replace("  ", " ")
    s = s.replace(" ", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def find_first_existing(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = set(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def to_datetime_series(s: pd.Series) -> pd.Series:
    if s is None:
        return s
    return pd.to_datetime(s.replace(list(ERROR_STRINGS), np.nan) if s.dtype == object else s, errors="coerce")


def to_numeric_series(s: pd.Series) -> pd.Series:
    if s is None:
        return s
    return pd.to_numeric(s.replace(list(ERROR_STRINGS), np.nan) if s.dtype == object else s, errors="coerce")


def first_non_null(df: pd.DataFrame, cols: List[str]) -> pd.Series:
    out = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    for c in cols:
        if c in df.columns:
            ser = pd.to_datetime(df[c], errors="coerce")
            out = out.fillna(ser)
    return out


def month_end(series: pd.Series) -> pd.Series:
    s = pd.to_datetime(series, errors="coerce")
    return s.dt.to_period("M").dt.to_timestamp("M")


def week_start_from_date(date_series: pd.Series, start_date: pd.Timestamp) -> pd.Series:
    d = pd.to_datetime(date_series, errors="coerce")
    delta_days = (d - start_date).dt.days
    week_no = np.floor(delta_days / 7).astype("Int64") + 1
    return week_no


def safe_sum(series: pd.Series) -> float:
    if series is None or len(series) == 0:
        return 0.0
    return float(pd.to_numeric(series, errors="coerce").fillna(0).sum())


# ============================================================
# EXCEL READING
# ============================================================
@st.cache_data(show_spinner=False)
def read_controls(excel_path: str) -> Dict[str, object]:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb["Controls"]
    return {
        "start_date": pd.to_datetime(ws["B3"].value),
        "weeks_horizon": int(ws["B4"].value or 13),
        "months_horizon": int(ws["B5"].value or 6),
        "fx_sgd_inr": float(ws["B6"].value or 72),
        "overdue_days": int(ws["B9"].value or 7),
        "large_exposure": float(ws["B10"].value or 5_000_000),
    }


@st.cache_data(show_spinner=False)
def read_opening_cash(excel_path: str, sheet_name: str) -> float:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    val = ws["B5"].value
    try:
        return float(val or 0)
    except Exception:
        return 0.0


@st.cache_data(show_spinner=False)
def extract_table(excel_path: str, sheet_name: str, table_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[sheet_name]
    if table_name not in ws.tables:
        raise ValueError(f"Table {table_name} not found in sheet {sheet_name}.")
    ref = ws.tables[table_name].ref
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = list(
        ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    if not rows:
        return pd.DataFrame()
    headers = [normalize_col(h) for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=headers)
    df = df.apply(lambda col: col.map(_clean_scalar))
    df.columns = [normalize_col(c) for c in df.columns]
    # drop fully empty rows
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def remove_repeated_header_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    cols = list(df.columns)
    # if first non-null values resemble headers, remove that row
    def is_header_like(row) -> bool:
        vals = [row.get(c) for c in cols]
        non_null = [v for v in vals if not _is_nullish(v)]
        if len(non_null) < 2:
            return False
        hits = 0
        for c, v in zip(cols, vals):
            if _is_nullish(v):
                continue
            sv = normalize_col(v)
            if sv == c:
                hits += 1
        return hits >= max(2, len(cols) // 2)

    mask = df.apply(is_header_like, axis=1)
    if mask.any():
        df = df.loc[~mask].copy()

    return df.reset_index(drop=True)


# ============================================================
# CLEANING / DERIVED COLUMNS
# ============================================================
def add_derived_columns(df: pd.DataFrame, entity: str, section: str, start_date: pd.Timestamp) -> pd.DataFrame:
    df = df.copy()

    # normalize common column names to a predictable set
    rename_map = {}
    for c in df.columns:
        rename_map[c] = normalize_col(c)
    df = df.rename(columns=rename_map)

    # coerce dates / numerics by column name patterns
    for c in df.columns:
        if any(k in c for k in ["date", "month"]):
            df[c] = pd.to_datetime(df[c], errors="coerce")
        elif any(k in c for k in ["amount", "value", "pct", "week_no", "week"]):
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # Identify likely date columns
    if section == "ar":
        cash_candidates = [
            "cash_receipt_date",
            "linked_expected_receipt_date",
            "manual_expected_receipt_date",
            "receipt_date",
            "invoice_date",
        ]
        pnl_candidates = ["invoice_date", "cash_receipt_date"]
        inflow_candidates = ["net_receipt_amount", "invoice_value"]
        scf_fee_col = "scf_fee_amount"
        scf_date_col = "scf_date"
    elif section == "ap":
        cash_candidates = ["cash_payment_date", "payment_date", "due_date", "invoice_date"]
        pnl_candidates = ["invoice_date", "due_date", "cash_payment_date"]
        inflow_candidates = ["amount_payable", "net_amount", "amount"]
        scf_fee_col = None
        scf_date_col = None
    else:  # opex
        cash_candidates = ["cash_payment_date", "payment_date", "due_date", "invoice_date"]
        pnl_candidates = ["invoice_date", "due_date", "cash_payment_date"]
        inflow_candidates = ["amount", "net_amount"]
        scf_fee_col = None
        scf_date_col = None

    df["entity"] = entity
    df["section"] = section.upper()

    # core dates are derived from source fields, not cached Excel formula cells
    df["pnl_date"] = first_non_null(df, pnl_candidates)
    df["invoice_month_end"] = month_end(df["invoice_date"]) if "invoice_date" in df.columns else pd.NaT
    if section == "ar":
        df["scf_month_end"] = month_end(df["scf_date"]) if "scf_date" in df.columns else pd.NaT

    # Specific amounts / formula equivalents
    if section == "ar":
        invoice_value = pd.to_numeric(df.get("invoice_value"), errors="coerce") if "invoice_value" in df.columns else pd.Series(np.nan, index=df.index)
        scf_pct = pd.to_numeric(df.get("scf_fee_pct"), errors="coerce") if "scf_fee_pct" in df.columns else pd.Series(0.0, index=df.index)
        scf_flag = df.get("scf", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()
        scf_yes = scf_flag.eq("yes")

        df["scf_fee_amount_calc"] = np.where(scf_yes, invoice_value * scf_pct, 0.0)
        if "scf_fee_amount" in df.columns:
            df["scf_fee_amount_clean"] = pd.to_numeric(df["scf_fee_amount"], errors="coerce").fillna(df["scf_fee_amount_calc"])
        else:
            df["scf_fee_amount_clean"] = df["scf_fee_amount_calc"]

        if "net_receipt_amount" in df.columns:
            df["net_receipt_amount_clean"] = pd.to_numeric(df["net_receipt_amount"], errors="coerce")
        else:
            df["net_receipt_amount_clean"] = np.nan
        df["net_receipt_amount_clean"] = df["net_receipt_amount_clean"].fillna(invoice_value - df["scf_fee_amount_clean"])

        # rebuild cash receipt date from formula logic
        manual = pd.to_datetime(df.get("manual_expected_receipt_date"), errors="coerce") if "manual_expected_receipt_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        linked = pd.to_datetime(df.get("linked_expected_receipt_date"), errors="coerce") if "linked_expected_receipt_date" in df.columns else manual
        receipt_date = pd.to_datetime(df.get("receipt_date"), errors="coerce") if "receipt_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        scf_date = pd.to_datetime(df.get("scf_date"), errors="coerce") if "scf_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        receipt_status = df.get("receipt_status", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()

        cash_date = linked.copy()
        cash_date = cash_date.where(~receipt_status.eq("received"), receipt_date)
        cash_date = cash_date.where(~scf_yes, scf_date)
        # final fallback
        cash_date = cash_date.fillna(linked).fillna(manual).fillna(receipt_date).fillna(scf_date)
        df["cash_date"] = cash_date
        df["week_no"] = np.floor((df["cash_date"] - start_date).dt.days / 7).astype("Int64") + 1

        df["scf_month_end"] = month_end(df["scf_date"]) if "scf_date" in df.columns else pd.NaT

    elif section == "ap":
        amount_payable = pd.to_numeric(df.get("amount_payable"), errors="coerce") if "amount_payable" in df.columns else pd.Series(np.nan, index=df.index)
        gst_itc = pd.to_numeric(df.get("gst_itc_amount"), errors="coerce") if "gst_itc_amount" in df.columns else pd.Series(0.0, index=df.index)
        net_amount = pd.to_numeric(df.get("net_amount"), errors="coerce") if "net_amount" in df.columns else pd.Series(np.nan, index=df.index)

        df["amount_payable_clean"] = amount_payable
        # if a net amount exists use it; otherwise derive it
        derived_net = (amount_payable - gst_itc).clip(lower=0)
        if net_amount.notna().any():
            df["net_amount_clean"] = net_amount.fillna(derived_net)
        else:
            df["net_amount_clean"] = derived_net
        df["amount_for_pnl"] = df["net_amount_clean"].fillna(amount_payable)

        payment_status = df.get("payment_status", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()
        payment_date = pd.to_datetime(df.get("payment_date"), errors="coerce") if "payment_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        due_date = pd.to_datetime(df.get("due_date"), errors="coerce") if "due_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        cash_date = due_date.copy()
        cash_date = cash_date.where(~payment_status.eq("paid"), payment_date)
        cash_date = cash_date.fillna(payment_date).fillna(due_date)
        df["cash_date"] = cash_date
        df["week_no"] = np.floor((df["cash_date"] - start_date).dt.days / 7).astype("Int64") + 1

    else:  # opex
        amount = pd.to_numeric(df.get("amount"), errors="coerce") if "amount" in df.columns else pd.Series(np.nan, index=df.index)
        gst_itc = pd.to_numeric(df.get("gst_itc_amount"), errors="coerce") if "gst_itc_amount" in df.columns else pd.Series(0.0, index=df.index)
        net_amount = pd.to_numeric(df.get("net_amount"), errors="coerce") if "net_amount" in df.columns else pd.Series(np.nan, index=df.index)

        df["amount_clean"] = amount
        derived_net = (amount - gst_itc).clip(lower=0)
        if net_amount.notna().any():
            df["net_amount_clean"] = net_amount.fillna(derived_net)
        else:
            df["net_amount_clean"] = derived_net
        df["amount_for_pnl"] = df["net_amount_clean"].fillna(amount)

        status = df.get("status", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().str.lower()
        payment_date = pd.to_datetime(df.get("payment_date"), errors="coerce") if "payment_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        due_date = pd.to_datetime(df.get("due_date"), errors="coerce") if "due_date" in df.columns else pd.Series(pd.NaT, index=df.index)
        cash_date = due_date.copy()
        cash_date = cash_date.where(~status.eq("paid"), payment_date)
        cash_date = cash_date.fillna(payment_date).fillna(due_date)
        df["cash_date"] = cash_date
        df["week_no"] = np.floor((df["cash_date"] - start_date).dt.days / 7).astype("Int64") + 1

    # clean week/month values from workbook if present
    if "week_no" in df.columns:
        df["week_no"] = pd.to_numeric(df["week_no"], errors="coerce").astype("Int64")
    if "invoice_month_end" in df.columns:
        df["invoice_month_end"] = pd.to_datetime(df["invoice_month_end"], errors="coerce")

    return df


def clean_section_table(df: pd.DataFrame, entity: str, section: str, start_date: pd.Timestamp) -> pd.DataFrame:
    if df.empty:
        return df
    df = remove_repeated_header_rows(df)
    df = add_derived_columns(df, entity=entity, section=section, start_date=start_date)

    # remove completely empty-like rows after derivation
    essential = {
        "ar": ["invoice_date", "invoice_value", "net_receipt_amount_clean", "cash_date"],
        "ap": ["invoice_date", "amount_payable_clean", "amount_for_pnl", "cash_date"],
        "opex": ["invoice_date", "amount_clean", "amount_for_pnl", "cash_date"],
    }[section]

    # keep rows that have at least one meaningful core value
    mask = pd.Series(False, index=df.index)
    for c in essential:
        if c in df.columns:
            mask = mask | df[c].notna()
    df = df.loc[mask].copy()

    # drop rows that are still just header text
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].replace(list(ERROR_STRINGS), np.nan)

    df = df.reset_index(drop=True)
    return df


# ============================================================
# FORECAST BUILDERS
# ============================================================
def build_weekly_forecast(
    ar: pd.DataFrame,
    ap: pd.DataFrame,
    opex: pd.DataFrame,
    opening_cash: float,
    start_date: pd.Timestamp,
    weeks_horizon: int,
    currency: str,
) -> pd.DataFrame:
    weeks = pd.DataFrame({"week_no": range(1, weeks_horizon + 1)})
    weeks["week_start"] = start_date + pd.to_timedelta((weeks["week_no"] - 1) * 7, unit="D")
    weeks["week_end"] = weeks["week_start"] + pd.to_timedelta(6, unit="D")

    ar_f = ar.copy()
    ap_f = ap.copy()
    opex_f = opex.copy()

    ar_f = ar_f.loc[ar_f["week_no"].between(1, weeks_horizon, inclusive="both")].copy()
    ap_f = ap_f.loc[ap_f["week_no"].between(1, weeks_horizon, inclusive="both")].copy()
    opex_f = opex_f.loc[opex_f["week_no"].between(1, weeks_horizon, inclusive="both")].copy()

    if ar_f.empty:
        ar_grp = pd.Series(dtype=float)
        scf_grp = pd.Series(dtype=float)
    else:
        ar_amt = ar_f["net_receipt_amount_clean"] if "net_receipt_amount_clean" in ar_f.columns else ar_f.get("base_amount", 0)
        ar_grp = ar_f.groupby("week_no")[ar_amt.name if hasattr(ar_amt, "name") else "net_receipt_amount_clean"].sum(min_count=1)
        if "scf_fee_amount_clean" in ar_f.columns:
            scf_grp = ar_f.groupby("week_no")["scf_fee_amount_clean"].sum(min_count=1)
        else:
            scf_grp = pd.Series(dtype=float)

    if ap_f.empty:
        ap_grp = pd.Series(dtype=float)
    else:
        ap_col = "amount_payable_clean" if "amount_payable_clean" in ap_f.columns else "amount_for_pnl"
        ap_grp = ap_f.groupby("week_no")[ap_col].sum(min_count=1)

    if opex_f.empty:
        opex_grp = pd.Series(dtype=float)
    else:
        op_col = "amount_clean" if "amount_clean" in opex_f.columns else "amount_for_pnl"
        opex_grp = opex_f.groupby("week_no")[op_col].sum(min_count=1)

    weeks["inflows"] = weeks["week_no"].map(ar_grp).fillna(0.0)
    weeks["ap_outflows"] = weeks["week_no"].map(ap_grp).fillna(0.0)
    weeks["opex_outflows"] = weeks["week_no"].map(opex_grp).fillna(0.0)
    weeks["scf_fee"] = weeks["week_no"].map(scf_grp).fillna(0.0)

    # for weekly cash, SCF fees are not cash outflows in the Excel top box;
    # the Excel top box uses AR net receipt and AP/OPEX outflows.
    weeks["outflows"] = weeks["ap_outflows"] + weeks["opex_outflows"]
    weeks["net"] = weeks["inflows"] - weeks["outflows"]
    weeks["closing_cash"] = opening_cash + weeks["net"].cumsum()

    weeks["week_label"] = weeks.apply(
        lambda r: f"W{int(r['week_no'])} ({r['week_start']:%d-%b} to {r['week_end']:%d-%b})",
        axis=1,
    )
    weeks["currency"] = currency
    weeks["opening_cash"] = opening_cash
    return weeks


def build_monthly_pnl(
    ar: pd.DataFrame,
    ap: pd.DataFrame,
    opex: pd.DataFrame,
    start_date: pd.Timestamp,
    months_horizon: int,
    currency: str,
) -> pd.DataFrame:
    month_ends = pd.date_range(start=start_date, periods=months_horizon, freq="ME")
    pnl = pd.DataFrame({"month_end": month_ends})
    pnl["month_label"] = pnl["month_end"].dt.strftime("%b-%Y")

    # AR revenue
    ar["invoice_month_end"] = pd.to_datetime(ar.get("invoice_month_end"), errors="coerce")
    ap["invoice_month_end"] = pd.to_datetime(ap.get("invoice_month_end"), errors="coerce")
    opex["invoice_month_end"] = pd.to_datetime(opex.get("invoice_month_end"), errors="coerce")

    rev = (
        ar.loc[ar["invoice_month_end"].isin(month_ends)]
        .groupby("invoice_month_end")["invoice_value" if "invoice_value" in ar.columns else "base_amount"]
        .sum(min_count=1)
    )

    cogs = pd.Series(dtype=float)
    if not ap.empty:
        ap_cogs = ap.copy()
        if "category" in ap_cogs.columns:
            ap_cogs = ap_cogs.loc[ap_cogs["category"].astype(str).str.upper().eq("COGS")]
        ap_amt_col = "amount_for_pnl" if "amount_for_pnl" in ap_cogs.columns else "amount_payable_clean"
        cogs = (
            ap_cogs.loc[ap_cogs["invoice_month_end"].isin(month_ends)]
            .groupby("invoice_month_end")[ap_amt_col]
            .sum(min_count=1)
        )

    opex_amt_col = "amount_for_pnl" if "amount_for_pnl" in opex.columns else "amount_clean"
    opex_sum = (
        opex.loc[opex["invoice_month_end"].isin(month_ends)]
        .groupby("invoice_month_end")[opex_amt_col]
        .sum(min_count=1)
    )

    scf = pd.Series(dtype=float)
    if "scf_month_end" in ar.columns:
        ar_scf = ar.copy()
        scf_col = "scf_fee_amount_clean" if "scf_fee_amount_clean" in ar_scf.columns else "scf_fee_amount"
        scf = (
            ar_scf.loc[ar_scf["scf_month_end"].isin(month_ends)]
            .groupby("scf_month_end")[scf_col]
            .sum(min_count=1)
        )

    pnl["revenue"] = pnl["month_end"].map(rev).fillna(0.0)
    pnl["cogs"] = pnl["month_end"].map(cogs).fillna(0.0)
    pnl["gross_profit"] = pnl["revenue"] - pnl["cogs"]
    pnl["opex"] = pnl["month_end"].map(opex_sum).fillna(0.0)
    pnl["scf_finance_cost"] = pnl["month_end"].map(scf).fillna(0.0)
    pnl["ebitda"] = pnl["gross_profit"] - pnl["opex"] - pnl["scf_finance_cost"]
    pnl["currency"] = currency
    return pnl


def build_overview_metrics(
    ar: pd.DataFrame,
    ap: pd.DataFrame,
    opex: pd.DataFrame,
    weekly: pd.DataFrame,
    pnl: pd.DataFrame,
    start_date: pd.Timestamp,
    overdue_days: int,
    large_exposure: float,
) -> Dict[str, float]:
    ar_open = ar.copy()
    # open AR: no receipt status received or cash date in the future
    if "receipt_status" in ar_open.columns:
        open_mask = ~ar_open["receipt_status"].astype(str).str.lower().eq("received")
    else:
        open_mask = ar_open["cash_date"].notna()
    overdue_mask = open_mask & ar_open["cash_date"].notna() & (ar_open["cash_date"] < (start_date - pd.Timedelta(days=overdue_days)))
    missing_ar_cash = ar_open["cash_date"].isna().sum()

    if "payment_status" in ap.columns:
        ap_missing = ap["cash_date"].isna().sum()
    else:
        ap_missing = ap["cash_date"].isna().sum()

    total_inflows = float(weekly["inflows"].sum())
    total_outflows = float(weekly["outflows"].sum())
    net_movement = float(weekly["net"].sum())
    closing_cash = float(weekly["closing_cash"].iloc[-1]) if not weekly.empty else 0.0
    negative_weeks = int((weekly["closing_cash"] < 0).sum())

    large_ar = float(
        ar.loc[
            pd.to_numeric(ar.get("net_receipt_amount_clean", ar.get("invoice_value", 0)), errors="coerce").fillna(0).abs() >= large_exposure
        ].shape[0]
    )
    large_ap = float(
        ap.loc[
            pd.to_numeric(ap.get("amount_for_pnl", ap.get("amount_payable_clean", 0)), errors="coerce").fillna(0).abs() >= large_exposure
        ].shape[0]
    )
    large_opex = float(
        opex.loc[
            pd.to_numeric(opex.get("amount_for_pnl", opex.get("amount_clean", 0)), errors="coerce").fillna(0).abs() >= large_exposure
        ].shape[0]
    )

    return {
        "total_inflows": total_inflows,
        "total_outflows": total_outflows,
        "net_movement": net_movement,
        "closing_cash": closing_cash,
        "negative_weeks": negative_weeks,
        "overdue_ar": int(overdue_mask.sum()),
        "missing_ar_cash": int(missing_ar_cash),
        "missing_ap_cash": int(ap_missing),
        "large_ar": int(large_ar),
        "large_ap": int(large_ap),
        "large_opex": int(large_opex),
    }


# ============================================================
# DASHBOARD RENDERING
# ============================================================
def format_money(x, currency):
    if pd.isna(x):
        return ""
    if currency == "INR":
        return f"₹{x:,.0f}"
    return f"S${x:,.2f}"


def show_kpis(metrics: Dict[str, float], currency: str):
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total inflows", format_money(metrics["total_inflows"], currency))
    c2.metric("Total outflows", format_money(metrics["total_outflows"], currency))
    c3.metric("Net movement", format_money(metrics["net_movement"], currency))
    c4.metric("Closing cash", format_money(metrics["closing_cash"], currency))

    c5, c6, c7, c8 = st.columns(4)
    c5.metric("Negative weeks", metrics["negative_weeks"])
    c6.metric("Overdue AR", metrics["overdue_ar"])
    c7.metric("Missing AR cash dates", metrics["missing_ar_cash"])
    c8.metric("Missing AP cash dates", metrics["missing_ap_cash"])


def plot_weekly(weekly: pd.DataFrame, currency: str):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=weekly["week_start"],
        y=weekly["closing_cash"],
        mode="lines+markers",
        name="Closing Cash",
    ))
    fig.add_trace(go.Bar(
        x=weekly["week_start"],
        y=weekly["net"],
        name="Net Movement",
        opacity=0.45,
    ))
    fig.update_layout(
        title="Weekly cash trend",
        xaxis_title="Week start",
        yaxis_title=f"Cash ({currency})",
        barmode="overlay",
        height=420,
        legend_orientation="h",
    )
    st.plotly_chart(fig, use_container_width=True)


def plot_monthly_pnl(pnl: pd.DataFrame, currency: str):
    fig = go.Figure()
    fig.add_trace(go.Bar(x=pnl["month_label"], y=pnl["revenue"], name="Revenue"))
    fig.add_trace(go.Bar(x=pnl["month_label"], y=pnl["cogs"], name="COGS"))
    fig.add_trace(go.Bar(x=pnl["month_label"], y=pnl["opex"], name="OPEX"))
    fig.add_trace(go.Bar(x=pnl["month_label"], y=pnl["scf_finance_cost"], name="SCF fee"))
    fig.add_trace(go.Scatter(x=pnl["month_label"], y=pnl["ebitda"], mode="lines+markers", name="EBITDA"))
    fig.update_layout(
        title="Monthly P&L projection",
        xaxis_title="Month",
        yaxis_title=f"Amount ({currency})",
        barmode="group",
        height=420,
        legend_orientation="h",
    )
    st.plotly_chart(fig, use_container_width=True)


def render_dashboard(entity: str, data: Dict[str, pd.DataFrame], controls: Dict[str, object], opening_cash: float):
    currency = ENTITY_TABLES[entity]["currency"]
    st.title(f"{entity} Dashboard")
    # 🔄 Refresh Button
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
    st.caption(f"Weekly cash forecast and monthly P&L projection in {currency}")

    show_kpis(data["metrics"], currency)

    st.subheader("Weekly cash forecast")
    weekly_disp = data["weekly"][["week_no", "week_start", "week_end", "inflows", "ap_outflows", "opex_outflows", "net", "closing_cash"]].copy()
    weekly_disp.columns = ["Week", "Week Start", "Week End", "Inflows", "AP Outflows", "OPEX Outflows", "Net", "Closing Cash"]
    st.dataframe(
        weekly_disp.style.format({
            "Inflows": lambda x: format_money(x, currency),
            "AP Outflows": lambda x: format_money(x, currency),
            "OPEX Outflows": lambda x: format_money(x, currency),
            "Net": lambda x: format_money(x, currency),
            "Closing Cash": lambda x: format_money(x, currency),
        }),
        use_container_width=True,
        height=320,
    )
    plot_weekly(data["weekly"], currency)

    st.subheader("Monthly P&L projection")
    pnl_disp = data["pnl"][["month_label", "revenue", "cogs", "gross_profit", "opex", "scf_finance_cost", "ebitda"]].copy()
    pnl_disp.columns = ["Month", "Revenue", "COGS", "Gross Profit", "OPEX", "SCF Finance Cost", "EBITDA"]
    st.dataframe(
        pnl_disp.style.format({
            "Revenue": lambda x: format_money(x, currency),
            "COGS": lambda x: format_money(x, currency),
            "Gross Profit": lambda x: format_money(x, currency),
            "OPEX": lambda x: format_money(x, currency),
            "SCF Finance Cost": lambda x: format_money(x, currency),
            "EBITDA": lambda x: format_money(x, currency),
        }),
        use_container_width=True,
        height=320,
    )
    plot_monthly_pnl(data["pnl"], currency)

    with st.expander("Cleaned source tables"):
        st.write("AR")
        st.dataframe(data["ar"], use_container_width=True, height=240)
        st.write("AP")
        st.dataframe(data["ap"], use_container_width=True, height=240)
        st.write("OPEX")
        st.dataframe(data["opex"], use_container_width=True, height=240)

    st.download_button(
        "Download weekly forecast CSV",
        data["weekly"].to_csv(index=False).encode("utf-8"),
        file_name=f"{entity}_weekly_forecast.csv",
        mime="text/csv",
    )
    st.download_button(
        "Download monthly P&L CSV",
        data["pnl"].to_csv(index=False).encode("utf-8"),
        file_name=f"{entity}_monthly_pnl.csv",
        mime="text/csv",
    )


# ============================================================
# DATA LOAD PIPELINE
# ============================================================
@st.cache_data(show_spinner=False)
def load_entity_dashboard(excel_path: str, entity: str) -> Dict[str, pd.DataFrame]:
    if entity not in ENTITY_TABLES:
        raise ValueError(f"Unknown entity: {entity}")

    cfg = ENTITY_TABLES[entity]
    controls = read_controls(excel_path)
    start_date = pd.to_datetime(controls["start_date"])
    opening_cash = read_opening_cash(excel_path, cfg["sheet"])

    ar_raw = extract_table(excel_path, cfg["sheet"], cfg["ar"])
    ap_raw = extract_table(excel_path, cfg["sheet"], cfg["ap"])
    opex_raw = extract_table(excel_path, cfg["sheet"], cfg["opex"])

    ar = clean_section_table(ar_raw, entity, "ar", start_date)
    ap = clean_section_table(ap_raw, entity, "ap", start_date)
    opex = clean_section_table(opex_raw, entity, "opex", start_date)

    # Limit to the current forecast horizon for the main outputs
    weeks_horizon = controls["weeks_horizon"]
    months_horizon = controls["months_horizon"]
    currency = cfg["currency"]

    weekly = build_weekly_forecast(ar, ap, opex, opening_cash, start_date, weeks_horizon, currency)
    pnl = build_monthly_pnl(ar, ap, opex, start_date, months_horizon, currency)
    metrics = build_overview_metrics(
        ar, ap, opex, weekly, pnl, start_date, controls["overdue_days"], controls["large_exposure"]
    )

    return {
        "ar": ar,
        "ap": ap,
        "opex": opex,
        "weekly": weekly,
        "pnl": pnl,
        "metrics": metrics,
        "controls": controls,
        "opening_cash": opening_cash,
    }


# ============================================================
# APP
# ============================================================
def main():
    st.set_page_config(page_title="MIS Dashboard", layout="wide")
    st.sidebar.title("Settings")

    uploaded = st.sidebar.file_uploader("Upload workbook", type=["xlsx"])
    if uploaded is not None:
        # Save to a temporary local file
        tmp_path = Path("/tmp/uploaded_mis.xlsx")
        tmp_path.write_bytes(uploaded.read())
        excel_path = str(tmp_path)
    else:
        excel_path = DEFAULT_FILE

    if not Path(excel_path).exists():
        st.info("Upload the Excel workbook or place it next to this app as 'Group MIS working file.v4.xlsx'.")
        return

    controls = read_controls(excel_path)
    st.sidebar.write("Start date:", controls["start_date"].date())
    st.sidebar.write("Weeks horizon:", controls["weeks_horizon"])
    st.sidebar.write("Months horizon:", controls["months_horizon"])
    st.sidebar.write("FX SGD/INR:", controls["fx_sgd_inr"])

    if ENTITY_FIXED in ENTITY_TABLES:
        entity = ENTITY_FIXED
    else:
        entity = st.sidebar.radio("Dashboard", ["A_INR", "E_SGD"], index=0)

    st.sidebar.write(f"Opening cash ({entity}): {read_opening_cash(excel_path, ENTITY_TABLES[entity]['sheet']):,.2f}")

    with st.spinner(f"Building {entity} dashboard..."):
        data = load_entity_dashboard(excel_path, entity)

    render_dashboard(entity, data, controls, data["opening_cash"])

    st.caption(
        "This app reads the Excel Tables directly from the workbook, cleans the AR/AP/OPEX sections, "
        "and rebuilds the forecast logic inside Streamlit."
    )


if __name__ == "__main__":
    main()
